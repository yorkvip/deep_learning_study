from openpyxl import load_workbook

path = r"d:\python\深度学习！\山东大学信息科学与工程学院学生综合素质测评表.xlsx"
wb = load_workbook(path)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} (max_row={ws.max_row}, max_col={ws.max_column}) ===")
    found = 0
    for row in ws.iter_rows():
        for cell in row:
            align = cell.alignment
            if align is None:
                continue
            tr = align.text_rotation
            if tr is not None and tr != 0:
                print(f"  {cell.coordinate}: value='{cell.value}', text_rotation={tr}")
                found += 1
    if found == 0:
        print("  No vertical/rotated text found.")
