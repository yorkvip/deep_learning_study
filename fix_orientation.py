from openpyxl import load_workbook
from openpyxl.styles import Alignment

path = r"d:\python\深度学习！\山东大学信息科学与工程学院学生综合素质测评表.xlsx"
wb = load_workbook(path)

count = 0
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in ws.iter_rows():
        for cell in row:
            align = cell.alignment
            if align is None:
                continue
            if align.text_rotation is not None and align.text_rotation != 0:
                new_align = Alignment(
                    horizontal=align.horizontal,
                    vertical=align.vertical,
                    text_rotation=0,
                    wrap_text=align.wrap_text,
                    shrink_to_fit=align.shrink_to_fit,
                    indent=align.indent,
                    relative_indent=align.relative_indent,
                    justify_last_line=align.justify_last_line,
                )
                cell.alignment = new_align
                count += 1
                print(f"  Fixed {sheet_name}!{cell.coordinate}: '{cell.value}' (was rotation={align.text_rotation})")

wb.save(path)
print(f"\nDone. Changed {count} cells to horizontal text.")
